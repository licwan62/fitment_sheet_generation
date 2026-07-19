import json

import pandas as pd
from openpyxl import load_workbook

from moto_dimension_crawler.exporter import build_fitment_rows, export_all


def test_fitment_view_prefers_complete_dimensions():
    inputs = [{"INPUT_ID":"000001","MAKE":"BMW","MODEL":"C400GT","车辆类型":"两轮摩托车"}]
    raw = [
        {"INPUT_ID":"000001","L-MM":2200,"W-MM":"","H-MM":1400,"MATCH_SCORE":95,"MATCH_CONFIDENCE":"HIGH","PARSE_STATUS":"PARTIAL","SOURCE_URL":"partial"},
        {"INPUT_ID":"000001","L-MM":2210,"W-MM":835,"H-MM":1437,"MATCH_SCORE":90,"MATCH_CONFIDENCE":"MEDIUM","PARSE_STATUS":"COMPLETE","SOURCE_URL":"complete"},
    ]
    row = build_fitment_rows(inputs, raw)[0]
    assert row["DIMENSION_STATUS"] == "FOUND_COMPLETE"
    assert row["SOURCE_URL"] == "complete"
    assert row["L-MM"] == 2210


def test_fitment_view_retains_not_found_input():
    row = build_fitment_rows([{"INPUT_ID":"000002","MAKE":"BMW","MODEL":"Unknown","车辆类型":""}], [])[0]
    assert row["DIMENSION_STATUS"] == "NOT_FOUND"
    assert row["L-MM"] is None


def test_fitment_view_marks_ai_dimensions_as_unverified_inference():
    inputs = [{"INPUT_ID":"000290","MAKE":"Honda","MODEL":"CB190R","车辆类型":""}]
    raw = [{
        "INPUT_ID":"000290", "L-MM":2029, "W-MM":739, "H-MM":1041,
        "MATCH_SCORE":0, "MATCH_CONFIDENCE":"LOW", "MATCH_STATUS":"INFERRED",
        "PARSE_STATUS":"COMPLETE", "DATA_SOURCE":"QWEN_INFERENCE", "SOURCE_URL":"",
    }]
    row = build_fitment_rows(inputs, raw)[0]
    assert row["DIMENSION_STATUS"] == "INFERRED_COMPLETE"
    assert row["DATA_SOURCE"] == "QWEN_INFERENCE"
    assert "manual verification" in row["NOTES"]


def test_export_only_keeps_required_tables_and_logs_details(tmp_path, monkeypatch):
    monkeypatch.delenv("MOTO_EXPORT_XLSX", raising=False)
    inputs = [{"INPUT_ID":"000002","MAKE":"BMW","MODEL":"Unknown","车辆类型":""}]
    candidates = [{
        "INPUT_ID":"000002","MAKE":"BMW","MODEL":"Unknown","CANDIDATE_TITLE":"BMW Other",
        "CANDIDATE_URL":"https://example.test/other","MATCH_SCORE":42,"MATCH_CONFIDENCE":"LOW",
        "MATCH_STATUS":"REVIEW","MATCH_REASON":"model_similarity=42",
    }, {
        "INPUT_ID":"000002","MAKE":"BMW","MODEL":"Unknown","CANDIDATE_TITLE":"BMW Unrelated",
        "CANDIDATE_URL":"https://example.test/unrelated","MATCH_SCORE":20,"MATCH_CONFIDENCE":"LOW",
        "MATCH_STATUS":"MODEL_MISMATCH","MATCH_REASON":"model_similarity=20",
    }]
    not_found = [{
        "INPUT_ID":"000002","MAKE":"BMW","MODEL":"Unknown","BEST_CANDIDATE":"",
        "BEST_SCORE":"","NOTES":"No credible candidate",
    }]
    export_all(tmp_path, inputs, candidates, [], [], [], not_found, {"input_count": 1})

    assert sorted(path.name for path in tmp_path.glob("*.csv")) == ["candidate_pages.csv", "dimensions_summary.csv"]
    candidate_frame = pd.read_csv(tmp_path / "candidate_pages.csv")
    assert "MATCH_REASON" not in candidate_frame.columns
    assert list(candidate_frame.columns) == ["INPUT_ID","MAKE","MODEL","DATA_SOURCE","SOURCE_PRIORITY","CANDIDATE_TITLE","CANDIDATE_URL","MATCH_SCORE","MATCH_CONFIDENCE","MATCH_STATUS"]
    assert "NOT_FOUND" in set(candidate_frame["MATCH_STATUS"])
    assert "MODEL_MISMATCH" not in set(candidate_frame["MATCH_STATUS"])
    assert "BMW Other" in set(candidate_frame["CANDIDATE_TITLE"])
    assert "BMW Unrelated" not in set(candidate_frame["CANDIDATE_TITLE"])

    log_rows = [json.loads(line) for line in (tmp_path / "logs" / "run_details.jsonl").read_text(encoding="utf-8").splitlines()]
    diagnostic = next(row for row in log_rows if row["record_type"] == "CANDIDATE_DIAGNOSTIC")
    assert diagnostic["payload"]["MATCH_REASON"] == "model_similarity=42"


def test_xlsx_contains_only_required_sheets(tmp_path, monkeypatch):
    monkeypatch.setenv("MOTO_EXPORT_XLSX", "1")
    export_all(tmp_path, [], [], [], [], [], [], {"input_count": 0})

    workbook = load_workbook(tmp_path / "motorcycle_dimensions.xlsx")
    assert workbook.sheetnames == ["CANDIDATE_PAGES", "DIMENSIONS_SUMMARY"]
    for sheet in workbook:
        assert sheet.freeze_panes == "A2"
        assert sheet["A1"].font.bold
