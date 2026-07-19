from __future__ import annotations

import json
from pathlib import Path
import os

import pandas as pd


INPUT_COLUMNS = ["INPUT_ID","MAKE","MODEL","车辆类型","MAKE_NORMALIZED","MODEL_NORMALIZED","MODEL_COMPACT","MODEL_NUMBER_TOKENS","MODEL_WORD_TOKENS"]
CANDIDATE_COLUMNS = ["INPUT_ID","MAKE","MODEL","DATA_SOURCE","SOURCE_PRIORITY","CANDIDATE_TITLE","CANDIDATE_URL","MATCH_SCORE","MATCH_CONFIDENCE","MATCH_STATUS"]
RAW_COLUMNS = ["INPUT_ID","MAKE","MODEL","车辆类型","DATA_SOURCE","SOURCE_PRIORITY","SOURCE_MAKE","SOURCE_MODEL","SOURCE_VERSION","YEAR","YEAR_START","YEAR_END","PAGE_TITLE","SOURCE_URL","L_RAW","W_RAW","H_RAW","UNIT_RAW","L-MM-MIN","L-MM-MAX","W-MM-MIN","W-MM-MAX","H-MM-MIN","H-MM-MAX","L-MM","W-MM","H-MM","WHEELBASE-MM","SEAT-HEIGHT-MM","GROUND-CLEARANCE-MM","WIDTH_SCOPE","HEIGHT_SCOPE","ACCESSORY_STATUS","DIMENSION_RAW","MODEL_SIMILARITY","MATCH_SCORE","MATCH_CONFIDENCE","MATCH_STATUS","PARSE_STATUS","CONFIDENCE","ANOMALY_FLAGS","NOTES","FETCHED_AT","CONTENT_HASH"]
SUMMARY_COLUMNS = ["DIMENSION_GROUP_ID","MAKE","MODEL","VERSION","YEARS","YEAR_START","YEAR_END","L-MM","W-MM","H-MM-MIN","H-MM-MAX","WIDTH_SCOPE","HEIGHT_SCOPE","ACCESSORY_STATUS","SOURCE_COUNT","DATA_SOURCES","SOURCE_URLS","CONFIDENCE","REVIEW_STATUS"]
REVIEW_COLUMNS = ["INPUT_ID","MAKE","MODEL","ISSUE_TYPE","CANDIDATE_URLS","RAW_VALUE","MATCH_SCORE","ANOMALY_FLAGS","RECOMMENDED_ACTION","NOTES"]
NOT_FOUND_COLUMNS = ["INPUT_ID","MAKE","MODEL","车辆类型","SEARCH_TERMS_USED","BEST_CANDIDATE","BEST_SCORE","NOTES"]
FITMENT_COLUMNS = ["INPUT_ID","MAKE","MODEL","车辆类型","DIMENSION_STATUS","L-MM","W-MM","H-MM","H-MM-MIN","H-MM-MAX","YEAR","MODEL_SIMILARITY","MATCH_SCORE","MATCH_CONFIDENCE","MATCH_STATUS","PARSE_STATUS","SOURCE_PAGE_COUNT","DATA_SOURCE","SOURCE_URL","NOTES"]

LEGACY_OUTPUT_FILES = [
    "input_normalized.csv",
    "fitment_dimensions.csv",
    "dimensions_raw.csv",
    "review_needed.csv",
    "not_found.csv",
    "run_report.json",
]


def build_fitment_rows(inputs: list[dict], raw: list[dict]) -> list[dict]:
    by_input: dict[str, list[dict]] = {}
    for row in raw:
        by_input.setdefault(row["INPUT_ID"], []).append(row)
    result = []
    confidence_rank = {"HIGH": 3, "MEDIUM": 2, "LOW": 1, "": 0}
    for item in inputs:
        sources = by_input.get(item["INPUT_ID"], [])
        if sources:
            def rank(row: dict) -> tuple:
                completeness = sum(row.get(key) not in (None, "") for key in ("L-MM", "W-MM", "H-MM"))
                return (completeness, row.get("PARSE_STATUS") == "COMPLETE", confidence_rank.get(row.get("MATCH_CONFIDENCE", ""), 0), float(row.get("MATCH_SCORE") or 0), -int(row.get("SOURCE_PRIORITY") or 99))
            best = max(sources, key=rank)
            present = sum(best.get(key) not in (None, "") for key in ("L-MM", "W-MM", "H-MM"))
            if best.get("MATCH_STATUS") == "INFERRED":
                status = "INFERRED_COMPLETE" if present == 3 else "INFERRED_PARTIAL" if present else "NOT_FOUND"
            else:
                status = "FOUND_COMPLETE" if present == 3 else "FOUND_PARTIAL" if present else "SOURCE_NO_DIMENSION"
        else:
            best, status = {}, "NOT_FOUND"
        result.append({
            "INPUT_ID": item["INPUT_ID"], "MAKE": item["MAKE"], "MODEL": item["MODEL"], "车辆类型": item.get("车辆类型", ""),
            "DIMENSION_STATUS": status, "L-MM": best.get("L-MM"), "W-MM": best.get("W-MM"), "H-MM": best.get("H-MM"),
            "H-MM-MIN": best.get("H-MM-MIN"), "H-MM-MAX": best.get("H-MM-MAX"), "YEAR": best.get("YEAR", ""),
            "MODEL_SIMILARITY": best.get("MODEL_SIMILARITY", ""), "MATCH_SCORE": best.get("MATCH_SCORE", ""),
            "MATCH_CONFIDENCE": best.get("MATCH_CONFIDENCE", ""), "MATCH_STATUS": best.get("MATCH_STATUS", ""),
            "PARSE_STATUS": best.get("PARSE_STATUS", ""), "SOURCE_PAGE_COUNT": len(sources),
            "DATA_SOURCE": best.get("DATA_SOURCE", ""), "SOURCE_URL": best.get("SOURCE_URL", ""),
            "NOTES": ("Unverified AI inference; manual verification required."
                      if best.get("MATCH_STATUS") == "INFERRED"
                      else "Best available source; see DIMENSIONS_RAW for all years."
                      if sources else "No automatically trusted source page."),
        })
    return result


def _candidate_table_rows(candidates: list[dict], not_found: list[dict]) -> list[dict]:
    by_input: dict[str, list[dict]] = {}
    for candidate in candidates:
        by_input.setdefault(candidate["INPUT_ID"], []).append(candidate)
    rows = []
    for input_candidates in by_input.values():
        credible = [row for row in input_candidates if row.get("MATCH_STATUS") in {"EXACT", "LIKELY", "MULTIPLE", "INFERRED"}]
        if credible:
            rows.extend(credible)
            continue
        review = [row for row in input_candidates if row.get("MATCH_STATUS") == "REVIEW"]
        if review:
            rows.append(max(review, key=lambda row: float(row.get("MATCH_SCORE") or 0)))
    existing_not_found = {row["INPUT_ID"] for row in rows if row.get("MATCH_STATUS") == "NOT_FOUND"}
    for item in not_found:
        if item["INPUT_ID"] in existing_not_found:
            continue
        rows.append({
            "INPUT_ID": item["INPUT_ID"], "MAKE": item["MAKE"], "MODEL": item["MODEL"],
            "CANDIDATE_TITLE": "NOT_FOUND", "CANDIDATE_URL": item.get("BEST_CANDIDATE", ""),
            "MATCH_SCORE": item.get("BEST_SCORE", ""), "MATCH_CONFIDENCE": "LOW",
            "MATCH_STATUS": "NOT_FOUND",
        })
    return rows


def _write_structured_log(output: Path, datasets: list[tuple[str, list[dict]]], report: dict) -> None:
    log_dir = output / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    with (log_dir / "run_details.jsonl").open("w", encoding="utf-8") as handle:
        for record_type, rows in datasets:
            for row in rows:
                handle.write(json.dumps({"record_type": record_type, "payload": row}, ensure_ascii=False, default=str) + "\n")
        handle.write(json.dumps({"record_type": "RUN_REPORT", "payload": report}, ensure_ascii=False, default=str) + "\n")


def export_all(output: Path, inputs: list[dict], candidates: list[dict], raw: list[dict], groups: list[dict], review: list[dict], not_found: list[dict], report: dict) -> None:
    output.mkdir(parents=True, exist_ok=True)
    candidate_rows = _candidate_table_rows(candidates, not_found)
    datasets = {
        "CANDIDATE_PAGES": (candidate_rows, CANDIDATE_COLUMNS, "candidate_pages.csv"),
        "DIMENSIONS_SUMMARY": (groups, SUMMARY_COLUMNS, "dimensions_summary.csv"),
    }
    for filename in LEGACY_OUTPUT_FILES:
        (output / filename).unlink(missing_ok=True)
    frames = {}
    for sheet, (rows, columns, filename) in datasets.items():
        frames[sheet] = pd.DataFrame(rows).reindex(columns=columns)
        frames[sheet].to_csv(output / filename, index=False, encoding="utf-8-sig")
    _write_structured_log(output, [
        ("INPUT_NORMALIZED", inputs),
        ("FITMENT_DIMENSIONS", build_fitment_rows(inputs, raw)),
        ("CANDIDATE_DIAGNOSTIC", candidates),
        ("DIMENSIONS_RAW", raw),
        ("REVIEW_NEEDED", review),
        ("NOT_FOUND", not_found),
    ], report)
    if os.environ.get("MOTO_EXPORT_XLSX") != "1":
        (output / "motorcycle_dimensions.xlsx").unlink(missing_ok=True)
        return
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    excel = output / "motorcycle_dimensions.xlsx"
    with pd.ExcelWriter(excel, engine="openpyxl") as writer:
        for sheet, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet, index=False)
    workbook = load_workbook(excel)
    for sheet in workbook:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 34
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index, cells in enumerate(sheet.iter_cols(), start=1):
            data_width = max((len(str(cell.value)) for cell in cells[1:] if cell.value is not None), default=0)
            header_width = min(len(str(cells[0].value or "")) + 2, 18)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(data_width + 2, header_width, 10), 40)
    workbook.save(excel)
